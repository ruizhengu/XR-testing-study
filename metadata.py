import re
from pathlib import Path

import bibtexparser
from openpyxl import Workbook, load_workbook
import os
import glob


class Metadata:
    def __init__(self):
        self.dl_dict = {
            "ACM": {"files": "acm*.bib", "sheet": "ACM"},
            "IEEE": {"files": "IEEE*.bib", "sheet": "IEEE"},
            "Scopus": {"files": "scopus*.bib", "sheet": "Scopus"},
            "ScienceDirect": {"files": "ScienceDirect*.bib", "sheet": "ScienceDirect"},
            "ProQuest": {"files": "ProQuest*.bib", "sheet": "ProQuest"}
        }

        # self.library = "ACM"
        # self.library = "IEEE"
        # self.library = "Scopus"
        # self.library = "ScienceDirect"
        self.library = "ProQuest"

        self.dl = self.dl_dict[self.library]
        self.iter_root = Path("iter-1")

        self.bib_folder = self.iter_root / "bib"
        self.bib_files = glob.glob(os.path.join(self.bib_folder, self.dl["files"]))
        self.excel_file = self.iter_root / "XR Testing Literature.xlsx"
        self.headers = [
            "Title",
            "Author",
            "Year",
            "Abstract",
            "Publisher",
            "DOI",
            "Booktitle",
            "Journal",
            "Keywords"
        ]
        self.keyword_frequencies = {}

        if os.path.exists(self.excel_file):
            self.workbook = load_workbook(self.excel_file)
            self.sheet = self.workbook[self.dl["sheet"]]
        else:
            self.workbook = Workbook()
            self.sheet = self.workbook.active
        self.sheet.append(self.headers)

    def run(self):
        if self.library == "ProQuest" or self.library == "Scopus":
            self.parse_bibtex_manual()
        else:
            self.parse_bibtex()

    def parse_bibtex(self):
        count_search_results = 0
        count_keyword_verification = 0
        for bib_file in self.bib_files:
            with open(bib_file, 'r', encoding='utf-8') as file:
                bib_data = bibtexparser.load(file)
            for entry in bib_data.entries:
                count_search_results += 1
                title = entry.get("title", "N/A")
                if self.keyword_verification(title):
                    count_keyword_verification += 1
                    self.save_metadata(entry)
                # else:
                #     print(title)
        print("count_search_results", count_search_results)
        print("count_keyword_verification", count_keyword_verification)

    def save_metadata(self, entry):
        title = entry.get("title", "N/A")
        author = entry.get("author", "N/A")
        year = entry.get("year", "N/A")
        abstract = entry.get("abstract", "N/A")
        publisher = entry.get("publisher", "N/A")
        doi = entry.get("doi", "N/A")
        booktitle = entry.get("booktitle", "N/A")
        journal = entry.get("journal", "N/A")
        keywords = entry.get("keywords", "N/A")
        self.sheet.append([
            title,
            author,
            year,
            abstract,
            publisher,
            doi,
            booktitle,
            journal,
            keywords
        ])
        self.workbook.save(self.excel_file)

    def parse_bibtex_manual(self):
        count_search_results = 0
        count_keyword_verification = 0
        for bib_file in self.bib_files:
            with open(bib_file, 'r', encoding='utf-8') as file:
                content = file.read()
            entries = re.split(r'(?=@\w+{)', content)
            entries = [entry for entry in entries if entry.strip()]
            for entry in entries:
                count_search_results += 1
                entry_dict = {}
                fields = ["title", "author", "year", "abstract", "publisher", "doi", "booktitle", "journal", "keywords"]
                for field in fields:
                    regex = r'\b' + field + r'\s*=\s*[{"]([^"}]+)[}"]'
                    match = re.search(regex, entry, re.IGNORECASE)
                    if match:
                        entry_dict[field] = match.group(1)
                    else:
                        entry_dict[field] = "N/A"
                title = entry_dict["title"]
                if self.keyword_verification(title):
                    count_keyword_verification += 1
                    self.save_metadata_manual(entry_dict)
        print("count_search_results", count_search_results)
        print("count_keyword_verification", count_keyword_verification)

    def save_metadata_manual(self, entry_dict):
        title = entry_dict["title"]
        author = entry_dict["author"]
        year = entry_dict["year"]
        abstract = entry_dict["abstract"]
        publisher = entry_dict["publisher"]
        doi = entry_dict["doi"]
        booktitle = entry_dict["booktitle"]
        journal = entry_dict["journal"]
        keywords = entry_dict["keywords"]
        self.sheet.append([
            title,
            author,
            year,
            abstract,
            publisher,
            doi,
            booktitle,
            journal,
            keywords
        ])
        self.workbook.save(self.excel_file)

    def keyword_verification(self, title):
        if ((
                re.search(r'virtual[-\s]reality', title, re.IGNORECASE) or
                re.search(r'augmented[-\s]reality', title, re.IGNORECASE) or
                re.search(r'mixed[-\s]reality', title, re.IGNORECASE) or
                re.search(r'extended[-\s]reality', title, re.IGNORECASE) or
                re.search(r'\bVR\b', title) or
                re.search(r'\bAR\b', title) or
                re.search(r'\bXR\b', title) or
                re.search(r'\bMR\b', title)
        ) and (
                re.search(r'test', title, re.IGNORECASE) or
                re.search(r'validat', title, re.IGNORECASE) or
                re.search(r'verif', title, re.IGNORECASE)
                # re.search(r'bug', title, re.IGNORECASE) or
                # re.search(r'defect', title, re.IGNORECASE) or
                # re.search(r'fault', title, re.IGNORECASE) or
                # re.search(r'error', title, re.IGNORECASE)
        )):
            # and not re.search(r'usability', title, re.IGNORECASE)):
            return True
        return False

    def update_keyword_frequency(self, title):
        title = title.lower()
        title = re.sub(r'[^\w\s]', '', title)
        words = title.split()

        for word in words:
            if word in self.keyword_frequencies:
                self.keyword_frequencies[word] += 1
            else:
                self.keyword_frequencies[word] = 1


if __name__ == '__main__':
    m = Metadata()
    m.run()
